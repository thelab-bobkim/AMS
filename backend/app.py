from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
from datetime import datetime, timedelta
from models import db, Employee, AttendanceRecord
from config import Config
from dauoffice_api import DauofficeAPIClient
import io, os, jwt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import calendar
from functools import wraps

app = Flask(__name__)
app.config.from_object(Config)
JWT_SECRET = os.environ.get('FLASK_SECRET_KEY', 'ams-jwt-secret-2026')
JWT_EXP_HOURS = 8

CORS(app)
db.init_app(app)

dauoffice_client = None
client_id = Config.get_client_id()
client_secret = Config.get_client_secret()
if client_id:
    dauoffice_client = DauofficeAPIClient(
        client_id=client_id, client_secret=client_secret,
        api_url=app.config.get('DAUOFFICE_API_URL', 'https://api.daouoffice.com')
    )

# ==================== 다우오피스 관리자 계정 (.env 기반) ====================
# .env에 등록된 계정 ID 목록 (htkim,hnsong,dijo)
ADMIN_USERS = [u.strip() for u in os.environ.get('ADMIN_USERS', 'htkim,hnsong,dijo').split(',') if u.strip()]
# 관리자 공통 비밀번호 (.env에 ADMIN_PASSWORD 설정, 미설정 시 기본값)
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'DstiAdmin2026!')

print(f"[App] 관리자 계정: {ADMIN_USERS}")

with app.app_context():
    db.create_all()


# ==================== JWT 인증 헬퍼 ====================

def make_token(user_id, username, is_admin, employee_id=None):
    payload = {
        'user_id':  user_id,
        'username': username,
        'is_admin': is_admin,
        'employee_id': employee_id,  # 일반 사용자 전용
        'exp': datetime.utcnow() + timedelta(hours=JWT_EXP_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm='HS256')

def decode_token(token):
    return jwt.decode(token, JWT_SECRET, algorithms=['HS256'])

def get_current_user():
    """헤더에서 JWT 파싱 → (user_id, username, is_admin) 반환, 실패 시 None"""
    auth = request.headers.get('Authorization', '')
    if not auth.startswith('Bearer '):
        return None
    try:
        return decode_token(auth[7:])
    except Exception:
        return None

def require_login(f):
    """로그인 필수 데코레이터"""
    @wraps(f)
    def wrapped(*args, **kwargs):
        user = get_current_user()
        if not user:
            return jsonify({'error': '로그인이 필요합니다.', 'code': 'UNAUTHORIZED'}), 401
        request.current_user = user
        return f(*args, **kwargs)
    return wrapped

def require_admin(f):
    """관리자 전용 데코레이터"""
    @wraps(f)
    def wrapped(*args, **kwargs):
        user = get_current_user()
        if not user:
            return jsonify({'error': '로그인이 필요합니다.', 'code': 'UNAUTHORIZED'}), 401
        if not user.get('is_admin'):
            return jsonify({'error': '관리자 권한이 필요합니다.', 'code': 'FORBIDDEN'}), 403
        request.current_user = user
        return f(*args, **kwargs)
    return wrapped


# ==================== 인증 API ====================

@app.route('/api/auth/login', methods=['POST'])
def login():
    """로그인 API
    - 관리자: 다우오피스 계정 ID (htkim/hnsong/dijo) + 관리자 비밀번호
    - 일반직원: 다우오피스 영문 ID (dauoffice_user_id) + 공통 비밀번호
                없으면 한글 이름으로도 로그인 가능 (수동 등록 직원 대비)
    """
    data = request.json or {}
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()

    if not username or not password:
        return jsonify({'error': '아이디와 비밀번호를 입력하세요.'}), 400

    # ── 관리자 로그인 ──────────────────────────────────────────
    if username in ADMIN_USERS:
        if password != ADMIN_PASSWORD:
            return jsonify({'error': '비밀번호가 틀렸습니다.'}), 401
        token = make_token(username, username, True)
        return jsonify({
            'token': token,
            'username': username,
            'display_name': username,
            'is_admin': True,
            'role': 'admin'
        })

    # ── 일반 직원 로그인 ───────────────────────────────────────
    # 1순위: 다우오피스 영문 ID (dauoffice_user_id) 로 조회
    emp = Employee.query.filter_by(dauoffice_user_id=username, is_active=True).first()

    # 2순위: 한글 이름으로 조회 (수동 등록 직원 대비)
    if not emp:
        emp = Employee.query.filter_by(name=username, is_active=True).first()

    if not emp:
        return jsonify({'error': f'"{username}" 계정을 찾을 수 없습니다.\n다우오피스 ID 또는 이름을 확인하세요.'}), 401

    # 비밀번호 검증: .env USER_DEFAULT_PASSWORD (미설정 시 기본값 1234)
    default_pw = os.environ.get('USER_DEFAULT_PASSWORD', '1234')
    # 빈값이면 기본값 1234 사용
    if not default_pw.strip():
        default_pw = '1234'
    if password != default_pw:
        return jsonify({'error': '비밀번호가 틀렸습니다.'}), 401

    token = make_token(emp.id, emp.name, False, employee_id=emp.id)
    return jsonify({
        'token': token,
        'username': emp.name,
        'display_name': emp.name,
        'dauoffice_id': emp.dauoffice_user_id or '',
        'is_admin': False,
        'role': 'user',
        'employee_id': emp.id
    })

@app.route('/api/auth/verify', methods=['GET'])
def verify_token():
    """토큰 유효성 확인"""
    user = get_current_user()
    if not user:
        return jsonify({'valid': False}), 401
    return jsonify({'valid': True, 'username': user.get('username'), 'is_admin': user.get('is_admin')})

@app.route('/api/auth/admins', methods=['GET'])
@require_admin
def list_admins():
    """관리자 목록 조회 (.env 기반)"""
    return jsonify([{'username': u, 'is_active': True} for u in ADMIN_USERS])


# ==================== 직원 관리 API ====================

@app.route('/api/employees', methods=['GET'])
@require_login
def get_employees():
    employees = Employee.query.filter_by(is_active=True).all()
    return jsonify([emp.to_dict() for emp in employees])

@app.route('/api/employees', methods=['POST'])
@require_admin
def create_employee():
    data = request.json
    employee = Employee(name=data['name'], department=data.get('department', ''), data_source='manual')
    db.session.add(employee)
    db.session.commit()
    return jsonify(employee.to_dict()), 201

@app.route('/api/employees/<int:employee_id>', methods=['PUT'])
@require_admin
def update_employee(employee_id):
    employee = Employee.query.get_or_404(employee_id)
    data = request.json
    employee.name = data.get('name', employee.name)
    employee.department = data.get('department', employee.department)
    db.session.commit()
    return jsonify(employee.to_dict())

@app.route('/api/employees/<int:employee_id>', methods=['DELETE'])
@require_admin
def delete_employee(employee_id):
    employee = Employee.query.get_or_404(employee_id)
    employee.is_active = False
    db.session.commit()
    return jsonify({'message': '직원이 삭제되었습니다.'})


# ==================== 출근 기록 API ====================

@app.route('/api/attendance', methods=['GET'])
@require_login
def get_attendance_records():
    """출근 기록 조회
    - 관리자: 전체 조회
    - 일반 사용자: 본인 이름과 일치하는 직원 기록만 반환
    """
    user = request.current_user
    year  = request.args.get('year',  type=int)
    month = request.args.get('month', type=int)
    department = request.args.get('department', '')

    if not year or not month:
        return jsonify({'error': '년도와 월을 지정해주세요.'}), 400

    start_date = datetime(year, month, 1).date()
    _, last_day = calendar.monthrange(year, month)
    end_date = datetime(year, month, last_day).date()

    emp_query = Employee.query.filter_by(is_active=True)

    # 🔐 일반 사용자: 본인 이름과 일치하는 직원만
    if not user.get('is_admin'):
        emp_query = emp_query.filter(Employee.name == user.get('username'))

    if department and user.get('is_admin'):
        emp_query = emp_query.filter_by(department=department)

    employees = emp_query.order_by(Employee.department, Employee.name).all()

    if not employees:
        return jsonify([])

    emp_ids = [e.id for e in employees]
    all_records = AttendanceRecord.query.filter(
        AttendanceRecord.employee_id.in_(emp_ids),
        AttendanceRecord.date >= start_date,
        AttendanceRecord.date <= end_date
    ).all()

    record_map = {}
    for rec in all_records:
        day_key = str(rec.date.day)
        try: cin = rec.check_in_time.strftime('%H:%M') if rec.check_in_time else ''
        except: cin = str(rec.check_in_time)[:5] if rec.check_in_time else ''
        record_map.setdefault(rec.employee_id, {})[day_key] = {
            'check_in': cin, 'type': (rec.record_type or 'normal'),
            'note': (rec.note or ''), 'source': (rec.data_source or 'manual'), 'id': rec.id
        }

    result = []
    for emp in employees:
        result.append({
            'id': emp.id, 'name': (emp.name or ''),
            'department': (emp.department or ''), 'days': record_map.get(emp.id, {})
        })
    return jsonify(result)

@app.route('/api/attendance', methods=['POST'])
@require_login
def create_attendance_record():
    """출근 기록 추가/수정
    - 관리자: 모든 직원 기록 수정 가능
    - 일반 사용자: 본인 기록만 수정 가능
    """
    user = request.current_user
    data = request.json
    employee_id = data['employee_id']

    # 🔐 권한 체크: 일반 사용자는 본인 기록만
    if not user.get('is_admin'):
        emp = Employee.query.get(employee_id)
        if not emp or emp.name != user.get('username'):
            return jsonify({'error': '본인의 기록만 수정할 수 있습니다.', 'code': 'FORBIDDEN'}), 403

    date_str = data['date']
    check_in_time_str = data.get('check_in_time')
    record_type = data.get('record_type', 'normal')
    note = data.get('note', '')

    date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
    check_in_time = None
    if check_in_time_str:
        check_in_time = datetime.strptime(check_in_time_str, '%H:%M:%S').time()

    existing = AttendanceRecord.query.filter_by(employee_id=employee_id, date=date_obj).first()
    if existing:
        if 'check_in_time' in (request.json or {}):
            existing.check_in_time = check_in_time
        existing.record_type = record_type
        existing.note = note
        existing.data_source = 'manual'
        db.session.commit()
        return jsonify(existing.to_dict())
    else:
        record = AttendanceRecord(
            employee_id=employee_id, date=date_obj,
            check_in_time=check_in_time, record_type=record_type,
            note=note, data_source='manual'
        )
        db.session.add(record)
        db.session.commit()
        return jsonify(record.to_dict()), 201

@app.route('/api/attendance/<int:record_id>', methods=['PUT'])
@require_login
def update_attendance_record(record_id):
    """출근 기록 수정
    - 관리자: 모든 기록 수정
    - 일반 사용자: 본인 기록만 수정
    """
    user = request.current_user
    record = AttendanceRecord.query.get_or_404(record_id)

    # 🔐 권한 체크
    if not user.get('is_admin'):
        emp = Employee.query.get(record.employee_id)
        if not emp or emp.name != user.get('username'):
            return jsonify({'error': '본인의 기록만 수정할 수 있습니다.', 'code': 'FORBIDDEN'}), 403

    data = request.json
    if 'check_in_time' in data and data['check_in_time']:
        record.check_in_time = datetime.strptime(data['check_in_time'], '%H:%M:%S').time()
    if 'record_type' in data: record.record_type = data['record_type']
    if 'note' in data: record.note = data['note']
    db.session.commit()
    return jsonify(record.to_dict())

@app.route('/api/attendance/<int:record_id>', methods=['DELETE'])
@require_login
def delete_attendance_record(record_id):
    """출근 기록 삭제
    - 관리자: 모든 기록 삭제
    - 일반 사용자: 본인 기록만 삭제
    """
    user = request.current_user
    record = AttendanceRecord.query.get_or_404(record_id)

    # 🔐 권한 체크
    if not user.get('is_admin'):
        emp = Employee.query.get(record.employee_id)
        if not emp or emp.name != user.get('username'):
            return jsonify({'error': '본인의 기록만 삭제할 수 있습니다.', 'code': 'FORBIDDEN'}), 403

    db.session.delete(record)
    db.session.commit()
    return jsonify({'message': '출근 기록이 삭제되었습니다.'})


# ==================== 다우오피스 연동 (관리자 전용) ====================

@app.route('/api/dauoffice/sync-employees', methods=['POST'])
@require_admin
def sync_employees():
    if not dauoffice_client:
        return jsonify({'error': '다우오피스 API가 설정되지 않았습니다.'}), 400
    try:
        count = dauoffice_client.sync_employees_from_dauoffice()
        return jsonify({'message': f'{count}명의 직원 정보가 동기화되었습니다.', 'count': count})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/dauoffice/sync-attendance', methods=['POST'])
@require_admin
def sync_attendance():
    if not dauoffice_client:
        return jsonify({'error': '다우오피스 API가 설정되지 않았습니다.'}), 400
    data = request.json
    year = data.get('year'); month = data.get('month')
    if not year or not month:
        return jsonify({'error': '년도와 월을 지정해주세요.'}), 400
    try:
        count = dauoffice_client.sync_attendance_from_dauoffice(year, month)
        return jsonify({'message': f'{count}개의 출근 기록이 동기화되었습니다.', 'count': count})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/excel', methods=['GET'])
@require_admin
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter
    import calendar
    from datetime import datetime, date as date_type
    import io

    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    if not year or not month:
        return jsonify({'error': '년도와 월을 지정해주세요.'}), 400

    department_filter = request.args.get('department', '')
    _, last_day = calendar.monthrange(year, month)

    query = Employee.query.filter_by(is_active=True)
    if department_filter:
        query = query.filter_by(department=department_filter)
    employees = query.order_by(Employee.department, Employee.name).all()

    start_date = date_type(year, month, 1)
    end_date   = date_type(year, month, last_day)
    emp_ids    = [e.id for e in employees]
    all_records = AttendanceRecord.query.filter(
        AttendanceRecord.employee_id.in_(emp_ids),
        AttendanceRecord.date >= start_date,
        AttendanceRecord.date <= end_date
    ).all()
    record_map = {}
    for rec in all_records:
        record_map.setdefault(rec.employee_id, {})[rec.date.day] = rec

    wb = Workbook(); ws = wb.active
    ws.title = f"{year}.{month:02d} 근태"

    C_TITLE_BG="1F4E79"; C_TITLE_FG="FFFFFF"; C_SAT_BG="D6E4F7"; C_SUN_BG="FDECEA"
    C_WEEK_BG="EBF3FB"; C_SUBHDR_BG="2E75B6"; C_LEAVE_BG="FFF2CC"; C_TRIP_BG="E2EFDA"
    C_TIME_FG="1F4E79"; C_LATE_BG="FCE4D6"; C_ODD_BG="F8FBFF"; C_EVEN_BG="FFFFFF"
    C_BORDER="B8CCE4"

    def fill(h): return PatternFill("solid", fgColor=h)
    def font(bold=False, color="000000", size=10, name="Malgun Gothic"):
        return Font(bold=bold, color=color, size=size, name=name)
    def center(wrap=False): return Alignment(horizontal='center', vertical='center', wrap_text=wrap)
    def border(color=C_BORDER, style='thin'):
        s = Side(style=style, color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    total_cols = 2 + last_day
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    tc = ws.cell(1, 1, value=f"{'['+department_filter+'] ' if department_filter else ''}{year}년 {month}월 출근 현황")
    tc.font=Font(bold=True, color=C_TITLE_FG, size=14, name="Malgun Gothic"); tc.fill=fill(C_TITLE_BG); tc.alignment=center(); ws.row_dimensions[1].height=32

    for ci, (lbl) in enumerate([("성  명",1), ("부서 / 팀",2)]):
        c = ws.cell(2, lbl[1], lbl[0]); c.font=font(True, C_TITLE_FG, 10); c.fill=fill(C_SUBHDR_BG); c.alignment=center(); c.border=border("4472C4","medium")

    for day in range(1, last_day+1):
        d = datetime(year, month, day); wd = d.weekday()
        wday_str = ['월','화','수','목','금','토','일'][wd]
        c = ws.cell(2, day+2, f"{day}\n({wday_str})")
        c.alignment=center(wrap=True); c.border=border(); c.font=font(True, size=9)
        if wd==5: c.fill=fill(C_SAT_BG); c.font=font(True,"2E75B6",9)
        elif wd==6: c.fill=fill(C_SUN_BG); c.font=font(True,"C00000",9)
        else: c.fill=fill(C_WEEK_BG)
    ws.row_dimensions[2].height=30
    ws.column_dimensions['A'].width=12; ws.column_dimensions['B'].width=16
    for day in range(1, last_day+1):
        ws.column_dimensions[get_column_letter(day+2)].width=9.5

    prev_dept=None; data_row=3
    for idx, emp in enumerate(employees):
        dept=emp.department or '미지정'; wd=data_row
        if dept!=prev_dept:
            if prev_dept is not None:
                ws.row_dimensions[wd].height=6
                for c in range(1, total_cols+1): ws.cell(wd,c).fill=fill("DAEEF3")
                data_row+=1; wd=data_row
            prev_dept=dept
        row_bg=C_ODD_BG if idx%2==0 else C_EVEN_BG
        nc=ws.cell(wd,1,emp.name); nc.font=font(True,size=10); nc.alignment=center(); nc.fill=fill(row_bg); nc.border=border()
        dc=ws.cell(wd,2,dept); dc.font=font(size=9,color="404040"); dc.alignment=center(); dc.fill=fill(row_bg); dc.border=border()
        for day in range(1, last_day+1):
            col=day+2; d=datetime(year,month,day); wd2=d.weekday()
            rec=record_map.get(emp.id,{}).get(day)
            cell_val=''; cell_fill=row_bg; cell_bold=False
            if wd2==5: cell_fill=C_SAT_BG
            elif wd2==6: cell_fill=C_SUN_BG
            if rec:
                rt=rec.record_type or 'normal'
                if rt=='annual_leave': cell_val='연 차'; cell_fill=C_LEAVE_BG; cell_bold=True
                elif rt=='half_leave': cell_val='반 차'; cell_fill=C_LEAVE_BG
                elif rt=='substitute_holiday': cell_val='대체휴무'; cell_fill="E2EFDA"
                elif rt=='business_trip': cell_val=f"출장\n{rec.note}" if rec.note else '출 장'; cell_fill=C_TRIP_BG
                elif rec.check_in_time:
                    try:
                        ci=rec.check_in_time; hh=ci.hour if hasattr(ci,'hour') else int(str(ci)[:2]); mm=ci.minute if hasattr(ci,'minute') else int(str(ci)[3:5])
                        cell_val=f"{hh:02d}:{mm:02d}"
                        if (hh,mm)>(9,30): cell_fill=C_LATE_BG; cell_bold=True
                        else: cell_fill=row_bg
                    except: cell_val=str(rec.check_in_time)[:5]
            c=ws.cell(wd,col,cell_val)
            c.font=Font(bold=cell_bold, color=C_TIME_FG if cell_val and ':' in cell_val else "000000", size=9, name="Malgun Gothic")
            c.fill=fill(cell_fill); c.alignment=center(wrap=True); c.border=border()
        ws.row_dimensions[wd].height=18; data_row+=1

    for c in range(1, total_cols+1):
        ws.cell(data_row,c).border=Border(top=Side(style='medium',color="4472C4"))
    ws.freeze_panes=ws.cell(3,3)
    from openpyxl.worksheet.page import PageMargins
    ws.page_setup.orientation=ws.ORIENTATION_LANDSCAPE; ws.page_setup.fitToPage=True
    ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
    ws.page_margins=PageMargins(left=0.5,right=0.5,top=0.75,bottom=0.75); ws.print_title_rows='1:2'

    output=io.BytesIO(); wb.save(output); output.seek(0)
    filename=(f"출근기록_{year}.{month:02d}{'_'+department_filter if department_filter else ''}.xlsx")
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)

@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({'status': 'ok', 'dauoffice_api_configured': dauoffice_client is not None})


# ==================== SenseLink 연동 (관리자 전용) ====================
import os as _os, requests as _req
SENSELINK_RELAY_URL = _os.environ.get('SENSELINK_RELAY_URL', 'http://175.198.93.89:8765')

@app.route('/api/senselink/sync', methods=['POST'])
@require_admin
def senselink_sync():
    data=request.json or {}; year=data.get('year',datetime.now().year); month=data.get('month',datetime.now().month)
    date_prefix=f"{year:04d}-{month:02d}-"; _,last_d=calendar.monthrange(year,month)
    HEADERS={'ngrok-skip-browser-warning':'true','User-Agent':'AMS/1.0'}
    all_recs,errors=[],[]; page,size=1,100; MAX_PAGES=200
    while page<=MAX_PAGES:
        try:
            resp=_req.get(f"{SENSELINK_RELAY_URL}/attendance",params={'page':page,'size':size,'year':year,'month':month,'start':f"{year:04d}-{month:02d}-01",'end':f"{year:04d}-{month:02d}-{last_d:02d}"},headers=HEADERS,timeout=30)
            if resp.status_code!=200: errors.append(f"relay HTTP {resp.status_code}"); break
            body=resp.json(); data_obj=body.get('data',{}); recs=data_obj.get('list',[]); total=data_obj.get('total',0)
            if not recs: break
            for rec in recs:
                st=rec.get('sign_time_str','')
                if st.startswith(date_prefix): all_recs.append(rec)
            oldest=min((r.get('sign_time_str','') for r in recs),default='')
            if oldest and oldest<f"{year:04d}-{month:02d}-01": break
            if page*size>=total: break
            page+=1
        except Exception as e: errors.append(str(e)); break
    best={}
    for rec in all_recs:
        uname=(rec.get('user_name') or '').strip(); stime=rec.get('sign_time_str','')
        if not uname or not stime or str(rec.get('type',''))!='1': continue
        dkey=stime[:10]; key=(uname,dkey); direction=str(rec.get('device_direction','0'))
        cur=best.get(key)
        if cur is None: best[key]=rec
        else:
            cur_dir=str(cur.get('device_direction','0'))
            if direction=='0' and cur_dir!='0': best[key]=rec
            elif direction==cur_dir and stime<cur.get('sign_time_str',''): best[key]=rec
    synced=0
    for (uname,dkey),rec in best.items():
        try:
            emp=Employee.query.filter_by(name=uname,is_active=True).first()
            if not emp: continue
            work_date=datetime.strptime(dkey,'%Y-%m-%d').date()
            stime=rec.get('sign_time_str',''); ci_time=None
            if len(stime)>=19: ci_time=datetime.strptime(stime[:19],'%Y-%m-%d %H:%M:%S').time()
            existing=AttendanceRecord.query.filter_by(employee_id=emp.id,date=work_date).first()
            if existing:
                if ci_time: existing.check_in_time=ci_time
                existing.data_source='senselink'
            else:
                db.session.add(AttendanceRecord(employee_id=emp.id,date=work_date,check_in_time=ci_time,record_type='normal',data_source='senselink'))
            synced+=1
        except: pass
    db.session.commit()
    return jsonify({'code':200,'message':f'{synced}건 동기화 완료','synced':synced,'fetched':len(all_recs),'errors':errors[:5],'relay_url':SENSELINK_RELAY_URL})

@app.route('/api/senselink/status', methods=['GET'])
@require_admin
def senselink_status():
    total=AttendanceRecord.query.filter_by(data_source='senselink').count()
    return jsonify({'code':200,'relay_url':SENSELINK_RELAY_URL,'total_senselink_records':total})

@app.route('/api/senselink/preview', methods=['GET'])
@require_admin
def senselink_preview():
    year=request.args.get('year',datetime.now().year,type=int); month=request.args.get('month',datetime.now().month,type=int)
    recs=AttendanceRecord.query.filter(AttendanceRecord.data_source=='senselink',db.extract('year',AttendanceRecord.date)==year,db.extract('month',AttendanceRecord.date)==month).limit(20).all()
    return jsonify({'code':200,'count':len(recs),'records':[r.to_dict() for r in recs]})

@app.route('/api/daou/sync/employees', methods=['POST'])
@require_admin
def daou_sync_employees_alias(): return sync_employees()

@app.route('/api/daou/sync/attendance', methods=['POST'])
@require_admin
def daou_sync_attendance_alias(): return sync_attendance()

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
